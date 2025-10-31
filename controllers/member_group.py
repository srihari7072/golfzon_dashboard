from odoo import http
from odoo.http import request, Response
import json
import io
import xlsxwriter
from datetime import datetime
from typing import Any, List, Dict, Optional
from urllib.parse import quote
import logging

_logger = logging.getLogger(__name__)


class MemberGroupController(http.Controller):

    # HELPER METHODS
    def format_date_field(self, date_value):
        """Format date fields - handles microseconds, timestamps, and strings"""
        if not date_value:
            return ""

        try:
            if isinstance(date_value, int):
                timestamp = date_value / 1_000_000.0
                date_obj = datetime.fromtimestamp(timestamp)
                return date_obj.strftime("%B %d, %Y")
            elif hasattr(date_value, "strftime"):
                return date_value.strftime("%B %d, %Y")
            elif isinstance(date_value, str):
                for fmt in ["%Y-%m-%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"]:
                    try:
                        return datetime.strptime(date_value, fmt).strftime("%B %d, %Y")
                    except ValueError:
                        continue
                return date_value
            return ""
        except Exception as e:
            _logger.warning(f"Date formatting error: {e}")
            return ""

    def _format_number(self, number):
        """Format number with thousand separators"""
        try:
            return "{:,}".format(int(number))
        except:
            return str(number)

    def _get_division_label(self, code):
        """Get human-readable division label from code"""
        divisions = {
            "condition": "Condition-based generation",
            "register": "Register member list",
            "manual": "Manual selection",
        }
        return divisions.get(code, code or "N/A")

    # 1. SEARCH MEMBER GROUPS
    @http.route("/golfzon/member_group/search", type="json", auth="user")
    def search_member_groups(self, start_date, end_date, group_name="", **kwargs):
        try:
            domain = [
                ("created_at", ">=", f"{start_date} 00:00:00"),
                ("created_at", "<=", f"{end_date} 23:59:59"),
            ]
            if group_name and group_name.strip():
                domain.append(("group_name", "ilike", group_name))

            GroupInfo = request.env["golf.group.info"].sudo()
            groups = GroupInfo.search(domain, order="created_at desc")

            data = []
            for group in groups:
                division = self._get_division_label(group.group_scd or group.state_scd)
                data.append({
                    "id": group.id,
                    "group_id": group.group_id,
                    "division": division,
                    "group_name": group.group_name or "N/A",
                    "member_count": self._format_number(group.member_count or 0),
                    "creation_date": group.created_at.strftime("%Y.%m.%d") if group.created_at else "",
                    "revision_date": group.updated_at.strftime("%Y.%m.%d") if group.updated_at else "",
                    "group_code": group.group_code or "",
                    "state_scd": group.state_scd or "",
                })

            return {"status": "success", "data": data, "count": len(data)}

        except Exception as e:
            _logger.error("search_member_groups error: %s", e, exc_info=True)
            return {"status": "error", "message": str(e), "data": []}

    # 2. DOWNLOAD EXCEL
    @http.route("/golfzon/member_group/download_excel", type="http", auth="user", csrf=False)
    def download_excel(self, group_id=None, **kwargs):
        try:
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {"in_memory": True})
            worksheet = workbook.add_worksheet("Group List")

            header_format = workbook.add_format({
                "bold": True, "bg_color": "#4A90E2", "color": "white",
                "align": "center", "valign": "vcenter", "border": 1, "font_size": 12
            })
            cell_format = workbook.add_format({"align": "left", "valign": "vcenter", "border": 1, "font_size": 11})
            cell_format_center = workbook.add_format({"align": "center", "valign": "vcenter", "border": 1, "font_size": 11})

            headers = ["Division", "Member Group Name", "Number of Members", "Creation Date", "Date of Revision"]
            worksheet.set_column(0, 0, 15); worksheet.set_column(1, 1, 30)
            worksheet.set_column(2, 2, 20); worksheet.set_column(3, 3, 18); worksheet.set_column(4, 4, 18)

            for col, h in enumerate(headers):
                worksheet.write(0, col, h, header_format)

            GroupInfo = request.env["golf.group.info"].sudo()
            groups = None
            group_name_for_file = "group_list"

            if group_id:
                try:
                    gid = int(group_id)
                    groups = GroupInfo.search([("id", "=", gid)]) or GroupInfo.search([("group_id", "=", gid)])
                    if groups:
                        group_name_for_file = groups[0].group_name or "group"
                except:
                    groups = GroupInfo.search([("group_code", "=", str(group_id))])
                    if groups:
                        group_name_for_file = groups[0].group_name or "group"
            else:
                groups = GroupInfo.search([], order="created_at desc", limit=100)
                group_name_for_file = "all_groups"

            if not groups:
                worksheet.merge_range(1, 0, 1, 4, "No data found", cell_format_center)
            else:
                row = 1
                for group in groups:
                    division_name = "-"
                    if hasattr(group, "division") and group.division:
                        division_name = str(group.division)
                    worksheet.write(row, 0, division_name, cell_format_center)
                    worksheet.write(row, 1, group.group_name or "-", cell_format)
                    worksheet.write(row, 2, str(group.member_count or 0), cell_format_center)
                    worksheet.write(row, 3, group.created_at.strftime("%Y.%m.%d") if group.created_at else "-", cell_format_center)
                    worksheet.write(row, 4, group.updated_at.strftime("%Y.%m.%d") if group.updated_at else "-", cell_format_center)
                    row += 1

            workbook.close()
            output.seek(0)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename_ascii = f"member_group_{timestamp}.xlsx"
            filename_utf8 = f"{group_name_for_file[:30]}_{timestamp}.xlsx"
            filename_encoded = quote(filename_utf8.encode("utf-8"))

            return request.make_response(
                output.getvalue(),
                headers=[
                    ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("Content-Disposition", f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{filename_encoded}"),
                    ("Content-Length", str(len(output.getvalue()))),
                ],
            )
        except Exception as e:
            _logger.error("Excel download error: %s", e, exc_info=True)
            return request.make_response(f"Error: {str(e)}", headers=[("Content-Type", "text/plain")], status=500)

    # 3. GET STATISTICS (GLOBAL)
    @http.route("/golfzon/member_group/get_statistics", type="json", auth="user")
    def get_statistics(self, **kwargs):
        try:
            GroupInfo = request.env["golf.group.info"].sudo()
            GroupMember = request.env["golf.group.member"].sudo()

            total_groups = GroupInfo.search_count([])
            total_members = GroupMember.search_count([])
            today = datetime.now()
            first_day = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            groups_this_month = GroupInfo.search_count([("created_at", ">=", first_day)])

            return {
                "status": "success",
                "data": {
                    "total_groups": total_groups,
                    "total_members": total_members,
                    "groups_this_month": groups_this_month,
                },
            }
        except Exception as e:
            _logger.error("get_statistics error: %s", e, exc_info=True)
            return {"status": "error", "message": str(e)}

    # 4. UPLOAD MEMBER LIST
    @http.route("/golfzon/member_group/upload_member_list", type="http", auth="user", methods=["POST"], csrf=False)
    def upload_member_list(self, group_title: Optional[str] = None, member_list_file: Optional[Any] = None, **kwargs):
        try:
            import openpyxl
            from io import BytesIO

            if not group_title or not member_list_file:
                return self._json_response({"status": "error", "message": "Group title and file required"})

            file_content = member_list_file.read()
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            sheet = workbook.active
            if not sheet:
                return self._json_response({"status": "error", "message": "Cannot read Excel"})

            required_headers = ["Membership number", "name", "Date of joining", "contact", "email", "Residence", "Number of times built-in", "Recent interior date"]
            actual_headers = [cell.value for cell in sheet[1]]
            missing = [h for h in required_headers if h not in actual_headers]
            if missing:
                return self._json_response({"status": "error", "message": f"Missing columns: {', '.join(missing)}"})

            members = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    members.append({
                        "membership_number": row[0], "name": row[1], "date_of_joining": row[2],
                        "contact": row[3], "email": row[4], "residence": row[5],
                        "built_in_times": row[6], "recent_date": row[7],
                    })

            if not members:
                return self._json_response({"status": "error", "message": "No member data"})

            GroupInfo = request.env["golf.group.info"].sudo()
            new_group = GroupInfo.create({
                "group_name": group_title,
                "member_count": len(members),
                "group_scd": "register",
                "state_scd": "active",
                "created_at": datetime.now(),
            })

            GroupMember = request.env["golf.group.member"].sudo()
            for m in members:
                GroupMember.create({
                    "group_id": new_group.id,
                    "person_code": m["membership_number"],
                    "group_member_name": m["name"],
                    "mobile_phone": m["contact"],
                    "email": m["email"],
                    "created_at": datetime.now(),
                })

            return self._json_response({
                "status": "success",
                "message": f"Registered {len(members)} members",
                "group_id": new_group.id,
            })

        except Exception as e:
            _logger.error("upload_member_list error: %s", e, exc_info=True)
            return self._json_response({"status": "error", "message": str(e)})

    def _json_response(self, data):
        return request.make_response(json.dumps(data), headers=[("Content-Type", "application/json")])

    # 5. DOWNLOAD TEMPLATE
    @http.route("/golfzon/member_group/download_template", type="http", auth="user", csrf=False)
    def download_template(self, **kwargs):
        try:
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {"in_memory": True})
            worksheet = workbook.add_worksheet("Member List")

            header_format = workbook.add_format({
                "bold": True, "bg_color": "#FFFFFF", "color": "#000000",
                "align": "center", "valign": "vcenter", "border": 1, "font_size": 11, "font_name": "Arial"
            })
            data_format = workbook.add_format({"align": "center", "valign": "vcenter", "border": 1, "font_size": 10})

            headers = ["Membership number", "name", "Date of joining", "contact", "email", "Residence", "Number of times built-in", "Recent interior date"]
            widths = [20, 15, 18, 18, 25, 18, 25, 22]
            for i, w in enumerate(widths):
                worksheet.set_column(i, i, w)
            for col, h in enumerate(headers):
                worksheet.write(0, col, h, header_format)

            sample_data = [
                ["M001", "John Doe", "2025-01-15", "010-1234-5678", "john@example.com", "Seoul", "10", "2025-03-01"],
                ["M002", "Jane Smith", "2025-02-20", "010-8765-4321", "jane@example.com", "Busan", "5", "2025-03-10"],
                ["M003", "김철수", "2025-03-05", "010-5555-6666", "kim@example.com", "Incheon", "8", "2025-03-15"],
            ]
            for r, row in enumerate(sample_data, 1):
                for c, val in enumerate(row):
                    worksheet.write(r, c, val, data_format)

            workbook.close()
            output.seek(0)
            return request.make_response(
                output.getvalue(),
                headers=[
                    ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                    ("Content-Disposition", "attachment; filename=member_template.xlsx"),
                ],
            )
        except Exception as e:
            _logger.error("download_template error: %s", e, exc_info=True)
            return request.make_response("Error", status=500)

    # 6. CONDITION INQUIRY (FILTERED GROUP + INDICATORS)

    @http.route("/golfzon/member_group/condition_inquiry", type="json", auth="user")
    def condition_inquiry(self, **kwargs):
        """
        Fetch detailed member info for selected group, supporting all filters.
        """
        try:
            cr = request.env.cr
            today = datetime.now()
            params = []

            # Get group id from name
            group_title = kwargs.get("group_title") or kwargs.get("group_name")
            if not group_title:
                return {
                    "status": "error",
                    "message": "Group title is required.",
                    "members": [],
                    "member_count": 0,
                    "number_of_members": 0,
                }
            
            cr.execute("SELECT id FROM golf_group_info WHERE group_name = %s LIMIT 1", (group_title,))
            row = cr.fetchone()
            if not row:
                return {
                    "status": "error",
                    "message": "Group not found.",
                    "members": [],
                    "member_count": 0,
                    "number_of_members": 0,
                }
            
            group_id = row[0]
            params.append(group_id)

            # Build filters
            where_conditions = ["gm.group_id = %s"]
            having_conditions = []

            # Gender filter
            gender = kwargs.get("gender")
            if gender and gender not in ("all", "selection", ""):
                where_conditions.append("gp.gender_scd = %s")
                params.append(gender)

            # Age group filter
            age_group = kwargs.get("age_group")
            if age_group and age_group != "unselected":
                year_now = today.year
                if age_group == "20s_under":
                    where_conditions.append("EXTRACT(YEAR FROM gp.birth_date) > %s")
                    params.append(year_now - 20)
                elif age_group == "20s":
                    where_conditions.append("EXTRACT(YEAR FROM gp.birth_date) BETWEEN %s AND %s")
                    params.extend([year_now - 29, year_now - 20])
                elif age_group == "30s":
                    where_conditions.append("EXTRACT(YEAR FROM gp.birth_date) BETWEEN %s AND %s")
                    params.extend([year_now - 39, year_now - 30])
                elif age_group == "40s":
                    where_conditions.append("EXTRACT(YEAR FROM gp.birth_date) BETWEEN %s AND %s")
                    params.extend([year_now - 49, year_now - 40])
                elif age_group == "50s_more":
                    where_conditions.append("EXTRACT(YEAR FROM gp.birth_date) <= %s")
                    params.append(year_now - 50)

            # Visit count filter
            visit_count = kwargs.get("visit_count")
            if visit_count and visit_count != "unselected":
                if visit_count == "5_under":
                    having_conditions.append("COUNT(vc.visit_seq) < 5")
                elif visit_count == "5_10":
                    having_conditions.append("COUNT(vc.visit_seq) BETWEEN 5 AND 10")
                elif visit_count == "10_more":
                    having_conditions.append("COUNT(vc.visit_seq) > 10")

            # Build WHERE and HAVING clauses
            where_clause = "WHERE " + " AND ".join(where_conditions) if where_conditions else ""
            having_clause = "HAVING " + " AND ".join(having_conditions) if having_conditions else ""

            # Compose query
            query = f"""
                SELECT
                    gm.group_member_name AS name,
                    mm.member_no AS membership_number,
                    mm.entry_date AS date_of_joining,
                    gp.mobile_phone AS contact,
                    gp.email,
                    gp.nation_scd AS residence,
                    COUNT(vc.visit_seq) AS number_of_times_builtin,
                    MAX(vc.visit_date) AS recent_interior_date
                FROM golf_group_member gm
                LEFT JOIN golfzon_person gp ON gp.person_code = gm.person_code
                LEFT JOIN members_members mm ON mm.person_code = gm.person_code
                LEFT JOIN visit_customers vc ON vc.person_code = gm.person_code
                {where_clause}
                GROUP BY gm.group_member_name, mm.member_no, mm.entry_date,
                        gp.mobile_phone, gp.email, gp.nation_scd
                {having_clause}
                ORDER BY gm.group_member_name ASC
            """

            cr.execute(query, tuple(params))
            members = []
            number_of_times_builtin_sum = 0  # Track total count
            for row in cr.fetchall():
                builtin_count = int(row[6] or 0)
                number_of_times_builtin_sum += builtin_count
                members.append({
                    "membership_number": row[1] or "",
                    "name": row[0] or "",
                    "date_of_joining": str(row[2]) if row[2] else "",
                    "contact": row[3] or "",
                    "email": row[4] or "",
                    "residence": row[5] or "",
                    "number_of_times_builtin": builtin_count,
                    "recent_interior_date": str(row[7]) if row[7] else "",
                })

            member_count = len(members)
            return {
                "status": "success",
                "member_count": member_count,
                "number_of_members": member_count,
                "indicators": {
                    "number_of_members": f"{member_count:,} people",
                    "number_of_times_builtin": f"{number_of_times_builtin_sum:,} times",  # <-- new dynamic value
                    # Add other indicators here as needed
                },
                "members": members,
            }
        except Exception as e:
            _logger = logging.getLogger(__name__)
            import traceback
            _logger.error("Error in condition_inquiry: %s\n%s", str(e), traceback.format_exc())
            return {
                "status": "error",
                "message": str(e),
                "members": [],
                "member_count": 0,
                "number_of_members": 0,
            }


        
    # 7. GET MEMBERS (PAGINATED)
    @http.route("/golfzon/member_group/get_members", type="json", auth="user")
    def get_members(self, group_id=None, offset=0, limit=10, **kwargs):
        try:
            query = """
                SELECT DISTINCT
                    gp.person_code, gp.member_name, mm.member_no, mm.entry_date,
                    gp.mobile_phone, gp.email, gp.nation_scd,
                    COALESCE(COUNT(vc.visit_seq), 0) AS visit_count,
                    MAX(vc.visit_date) AS last_visit
                FROM golfzon_person gp
                LEFT JOIN members_members mm ON mm.person_code = gp.person_code AND mm.deleted_at IS NULL
                LEFT JOIN visit_customers vc ON vc.person_code = gp.person_code AND vc.deleted_at IS NULL
                WHERE gp.deleted_at IS NULL
            """
            params = []
            if group_id:
                query += " AND EXISTS (SELECT 1 FROM golf_group_member gm WHERE gm.person_code = gp.person_code AND gm.group_id = %s)"
                params.append(int(group_id))

            query += """
                GROUP BY gp.person_code, gp.member_name, mm.member_no, mm.entry_date, gp.mobile_phone, gp.email, gp.nation_scd
                ORDER BY last_visit DESC NULLS LAST
                LIMIT %s OFFSET %s
            """
            params.extend([limit, offset])

            request.cr.execute(query, tuple(params))
            records = request.cr.dictfetchall()

            # Total count
            count_query = "SELECT COUNT(DISTINCT gp.person_code) FROM golfzon_person gp WHERE gp.deleted_at IS NULL"
            count_params = []
            if group_id:
                count_query += " AND EXISTS (SELECT 1 FROM golf_group_member gm WHERE gm.person_code = gp.person_code AND gm.group_id = %s)"
                count_params.append(int(group_id))
            request.cr.execute(count_query, tuple(count_params))
            total = request.cr.fetchone()[0] or 0

            members = []
            for r in records:
                members.append({
                    "membership_number": r["member_no"] or "",
                    "name": r["member_name"] or "",
                    "date_of_joining": self.format_date_field(r["entry_date"]),
                    "contact": r["mobile_phone"] or "",
                    "email": r["email"] or "",
                    "residence": r["nation_scd"] or "",
                    "number_of_times_builtin": str(r["visit_count"]),
                    "recent_interior_date": self.format_date_field(r["last_visit"]),
                })

            return {
                "status": "success",
                "data": members,
                "count": len(members),
                "total": total,
                "has_more": (offset + limit) < total,
                "offset": offset,
                "limit": limit,
            }

        except Exception as e:
            _logger.error("get_members error: %s", e, exc_info=True)
            return {"status": "error", "message": str(e), "data": [], "count": 0, "total": 0, "has_more": False}

    # 8. CALCULATE KEY INDICATORS
    def _calculate_key_indicators(self, member_records: List[Dict]) -> Dict[str, Any]:
        try:
            count = len(member_records)
            if not count:
                return {
                    "number_of_members": "0 people",
                    "number_of_times_builtin": "0 times",
                    "total_sales": "0 won",
                    "payment_green_fee": "0 won",
                    "open_green_fee": "0 won",
                    "average_discount_rate": "0%",
                }

            person_codes = tuple([r["person_code"] for r in member_records])
            query = """
                SELECT
                    COUNT(*) as usage_count,
                    COALESCE(SUM(total_amount), 0) as total_sales,
                    COALESCE(SUM(CASE WHEN fee_type = 'green_fee' THEN amount ELSE 0 END), 0) as green_fee_payment,
                    COALESCE(SUM(CASE WHEN fee_type = 'open_green_fee' THEN amount ELSE 0 END), 0) as open_green_fee,
                    COALESCE(AVG(discount_rate), 0) as avg_discount_rate
                FROM booking_bookings
                WHERE person_code IN %s AND deleted_at IS NULL
            """
            request.cr.execute(query, (person_codes,))
            data = request.cr.dictfetchone() or {}

            return {
                "number_of_members": f"{count:,} people",
                "number_of_times_builtin": f"{data.get('usage_count', 0):,} times",
                "total_sales": f"{int(data.get('total_sales', 0)):,} won",
                "payment_green_fee": f"{int(data.get('green_fee_payment', 0)):,} won",
                "open_green_fee": f"{int(data.get('open_green_fee', 0)):,} won",
                "average_discount_rate": f"{float(data.get('avg_discount_rate', 0)):.1f}%",
            }
        except Exception as e:
            _logger.warning("Indicator calculation failed: %s", e)
            return {
                "number_of_members": "0 people",
                "number_of_times_builtin": "0 times",
                "total_sales": "0 won",
                "payment_green_fee": "0 won",
                "open_green_fee": "0 won",
                "average_discount_rate": "0%",
            }